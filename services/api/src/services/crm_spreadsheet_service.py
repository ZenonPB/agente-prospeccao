"""CrmSpreadsheetService — preenche a planilha de CRM a partir dos leads do sistema.

Fluxo (docs/planilha-atualizar.md + requisito real do dono):
1. O usuário envia o arquivo `Planilha_aprimorada.xlsx`.
2. O sistema lista os leads do banco atribuídos ao consultor (carteira).
3. Para cada lead monta uma linha no formato exato da aba do consultor.
4. Preenche a partir da primeira linha livre após o cabeçalho e devolve o arquivo.

Regra da coluna LEAD: se o lead tiver contato primário (is_primary) → nome da
pessoa; senão → nome da empresa.
"""
import logging
import os
import re
import sys
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "..", "workers", "src"))
from src.db.models import (  # noqa: E402
    Contact, FollowUp, FollowUpStep, Lead, LeadStatus,
)

logger = logging.getLogger(__name__)

# Colunas da planilha por índice (1-based) — espelham a aba dos consultores.
COLS = {
    "LEAD": 1, "EMPRESA": 2, "PROSPECCAO": 3, "PITCH_ENVIADO": 4, "PITCH": 5,
    "FU1": 6, "FU2": 7, "FU3": 8, "RESPONDEU": 9, "CARGO": 10, "OBS": 11,
}

RESPONDEU_STATUSES = {
    LeadStatus.RESPONDIDO, LeadStatus.REUNIAO_MARCADA, LeadStatus.REUNIAO_FEITA,
    LeadStatus.PROPOSTA_ENVIADA,
}
PITCH_ENVIADO_STATUSES = {
    LeadStatus.CONTATADO, LeadStatus.RESPONDIDO, LeadStatus.REUNIAO_MARCADA,
    LeadStatus.REUNIAO_FEITA, LeadStatus.PROPOSTA_ENVIADA, LeadStatus.PERDIDO,
}
# Variação por string .value — comparação robusta com objetos ORM/Namespace.
RESPONDEU_VALUES = {s.value for s in RESPONDEU_STATUSES}
PITCH_ENVIADO_VALUES = {s.value for s in PITCH_ENVIADO_STATUSES}


def _dt_to_excel_datetime(value: Optional[datetime]) -> Optional[datetime]:
    """Normaliza para datetime puro (sem timezone) para gravar na célula."""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return value


def build_planilha_row(
    lead: Lead,
    primary_contact: Optional[Contact],
    followups: Dict[FollowUpStep, datetime],
) -> List[Any]:
    """Monta a linha da planilha para um lead (ordem das colunas da aba)."""
    if primary_contact and primary_contact.name:
        lead_name = primary_contact.name
        cargo = primary_contact.role_label
    else:
        lead_name = lead.company_name or lead.name or ""
        cargo = None

    prospeccao = lead.assigned_at or lead.created_at
    pitch = followups.get(FollowUpStep.OPENING) or lead.last_contacted_at
    fu1 = followups.get(FollowUpStep.FOLLOWUP_1)
    fu2 = followups.get(FollowUpStep.FOLLOWUP_2)
    fu3 = followups.get(FollowUpStep.CLOSING)

    status = lead.status
    status_val = status.value if status else None
    pitch_enviado = status_val in PITCH_ENVIADO_VALUES or bool(pitch)
    respondeu = "SIM" if status_val in RESPONDEU_VALUES else ("NÃO" if pitch_enviado else None)

    row = [""] * (max(COLS.values()))
    row[COLS["LEAD"] - 1] = lead_name
    row[COLS["EMPRESA"] - 1] = lead.company_name or lead.name or ""
    row[COLS["PROSPECCAO"] - 1] = _dt_to_excel_datetime(prospeccao)
    row[COLS["PITCH_ENVIADO"] - 1] = bool(pitch_enviado)
    row[COLS["PITCH"] - 1] = _dt_to_excel_datetime(pitch)
    row[COLS["FU1"] - 1] = _dt_to_excel_datetime(fu1)
    row[COLS["FU2"] - 1] = _dt_to_excel_datetime(fu2)
    row[COLS["FU3"] - 1] = _dt_to_excel_datetime(fu3)
    row[COLS["RESPONDEU"] - 1] = respondeu
    row[COLS["CARGO"] - 1] = cargo
    row[COLS["OBS"] - 1] = lead.notes
    return row


def _normalize(text: Any) -> str:
    """Minúsculas sem espaços extras — para comparação de duplicatas."""
    if text is None:
        return ""
    return re.sub(r"\s+", "", str(text).lower())


def complementa_planilha(
    workbook_path: str,
    sheet_name: str,
    leads: List[Lead],
    contacts_by_lead: Dict[str, List[Contact]],
    followups_by_lead: Dict[str, Dict[FollowUpStep, datetime]],
) -> Dict[str, int]:
    """Abre o xlsx, localiza a aba e anexa os leads na 1ª linha livre.

    Retorna {inseridos, duplicados} e salva o arquivo no mesmo caminho.
    """
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada na planilha")
    ws = wb[sheet_name]

    # Duplicatas já presentes na aba (col LEAD + Empresa).
    existing = set()
    for r in range(2, ws.max_row + 1):
        lead_cell = ws.cell(r, COLS["LEAD"]).value
        emp_cell = ws.cell(r, COLS["EMPRESA"]).value
        if lead_cell or emp_cell:
            existing.add((_normalize(lead_cell), _normalize(emp_cell)))

    inserted = 0
    duplicates = 0
    for lead in leads:
        contacts = contacts_by_lead.get(str(lead.id), [])
        primary = next((c for c in contacts if c.is_primary), contacts[0] if contacts else None)
        followups = followups_by_lead.get(str(lead.id), {})
        row = build_planilha_row(lead, primary, followups)

        lead_name = row[COLS["LEAD"] - 1]
        empresa = row[COLS["EMPRESA"] - 1]
        if (_normalize(lead_name), _normalize(empresa)) in existing:
            duplicates += 1
            continue
        existing.add((_normalize(lead_name), _normalize(empresa)))

        target = None
        for r in range(2, ws.max_row + 1):
            if not ws.cell(r, COLS["LEAD"]).value and not ws.cell(r, COLS["EMPRESA"]).value:
                target = r
                break
        if target is None:
            target = ws.max_row + 1

        for col, value in enumerate(row, start=1):
            if value is not None and value != "":
                ws.cell(target, col, value)
        inserted += 1

    wb.save(workbook_path)
    return {"inseridos": inserted, "duplicados": duplicates}