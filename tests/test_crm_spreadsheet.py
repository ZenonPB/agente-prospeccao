"""Testes do serviço de preenchimento da planilha de CRM.

Foca na lógica pura: `build_planilha_row` (mapeamento LEAD/pitch/follow-ups)
e a regra de LEAD (contato primário vs empresa). Sem banco real — usa objetos
simples com os atributos usados.
"""
from datetime import datetime, timezone
from types import SimpleNamespace

from src.services.crm_spreadsheet_service import (
    PITCH_ENVIADO_STATUSES,
    RESPONDEU_STATUSES,
    build_planilha_row,
)


def _lead(status="NOVO", company="Clinica Maua Ltda", assigned=None, last_contacted=None, notes=None):
    return SimpleNamespace(
        id="lead-1", company_name=company, name="Clinica Maua",
        assigned_at=assigned or datetime(2026, 7, 29, tzinfo=timezone.utc),
        created_at=datetime(2026, 7, 29, tzinfo=timezone.utc),
        last_contacted_at=last_contacted,
        notes=notes,
        status=SimpleNamespace(value=status),
    )


def _contact(name="Fabio Prada Perez", role="CEO", is_primary=True):
    return SimpleNamespace(name=name, role_label=role, is_primary=is_primary)


def test_lead_usa_contato_primario():
    row = build_planilha_row(
        _lead(status="NOVO"), _contact(), {},
    )
    # Col LEAD (1) = contato, Empresa (2) = empresa, CARGO (10) = CEO.
    assert row[0] == "Fabio Prada Perez"
    assert row[1] == "Clinica Maua Ltda"
    assert row[9] == "CEO"


def test_lead_sem_contato_usa_empresa():
    row = build_planilha_row(_lead(status="NOVO"), None, {})
    assert row[0] == "Clinica Maua Ltda"
    assert row[9] is None


def test_respondeu_sim_para_status_validos():
    for status in RESPONDEU_STATUSES:
        row = build_planilha_row(_lead(status=status.value), None, {})
        assert row[8] == "SIM"


def test_pitch_enviado_boolean():
    row = build_planilha_row(_lead(status="CONTATADO"), None, {})
    assert row[3] is True
    # Sem pitch e status NOVO → False.
    row = build_planilha_row(_lead(status="NOVO"), None, {})
    assert row[3] is False


def test_followups_vem_do_mapa():
    fu1 = datetime(2026, 8, 7, tzinfo=timezone.utc)
    fu2 = datetime(2026, 8, 10, tzinfo=timezone.utc)
    from src.db.models import FollowUpStep
    row = build_planilha_row(_lead(status="CONTATADO", last_contacted=datetime(2026, 8, 3, tzinfo=timezone.utc)), None, {
        FollowUpStep.FOLLOWUP_1: fu1,
        FollowUpStep.FOLLOWUP_2: fu2,
    })
    # Col FU1 (6) e FU2 (7). PITCH (5) usa last_contacted_at.
    assert row[5] == fu1.replace(tzinfo=None)
    assert row[6] == fu2.replace(tzinfo=None)
    assert row[4].year == 2026 and row[4].month == 8 and row[4].day == 3


# ---- Testes de complementa_planilha com criação de aba ----

import os
import tempfile

from openpyxl import Workbook

from src.services.crm_spreadsheet_service import complementa_planilha


def _make_xlsx(tmp_dir: str, sheets: dict[str, list[str]]) -> str:
    """Cria um .xlsx temporário com as abas e headers indicados.

    ``sheets`` mapeia nome_da_aba → lista de valores da primeira linha (header).
    """
    wb = Workbook()
    # remove a aba padrão criada pelo openpyxl
    default = wb.active
    for name, headers in sheets.items():
        ws = wb.create_sheet(title=name)
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, h)
    # remove aba "Sheet" se existir e tiver sobrado
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    path = os.path.join(tmp_dir, "test.xlsx")
    wb.save(path)
    wb.close()
    return path


def test_complementa_aba_existente(tmp_path):
    path = _make_xlsx(str(tmp_path), {"Consultor A": ["LEAD", "Empresa"]})
    result = complementa_planilha(path, "Consultor A", [], {}, {})
    assert result == {"inseridos": 0, "duplicados": 0}


def test_complementa_aba_inexistente_levanta_erro(tmp_path):
    path = _make_xlsx(str(tmp_path), {"Consultor A": ["LEAD", "Empresa"]})
    try:
        complementa_planilha(path, "NaoExiste", [], {}, {})
        assert False, "Deveria ter levantado ValueError"
    except ValueError as e:
        assert "não encontrada" in str(e)


def test_complementa_criar_aba_nova(tmp_path):
    path = _make_xlsx(str(tmp_path), {"Consultor A": ["LEAD", "Empresa", "Prospecção"]})
    result = complementa_planilha(
        path, "Nova Aba", [], {}, {},
        criar_aba_se_ausente=True,
    )
    assert result == {"inseridos": 0, "duplicados": 0}
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert "Nova Aba" in wb.sheetnames
    ws = wb["Nova Aba"]
    assert ws.cell(1, 1).value == "LEAD"
    assert ws.cell(1, 2).value == "Empresa"
    assert ws.cell(1, 3).value == "Prospecção"
    wb.close()


def test_complementa_criar_aba_insere_leads(tmp_path):
    path = _make_xlsx(str(tmp_path), {"Minha Aba": ["LEAD", "Empresa", "Prospecção"]})
    lead = SimpleNamespace(
        id="l1", company_name="Acme Corp", name="Acme",
        assigned_at=datetime(2026, 8, 1, tzinfo=timezone.utc),
        created_at=datetime(2026, 8, 1, tzinfo=timezone.utc),
        last_contacted_at=None, notes="nota teste",
        status=SimpleNamespace(value="NOVO"),
    )
    contact = SimpleNamespace(name="João Silva", role_label="Gerente", is_primary=True)
    contacts = {"l1": [contact]}
    result = complementa_planilha(
        path, "Minha Aba", [lead], contacts, {},
        criar_aba_se_ausente=True,
    )
    assert result["inseridos"] == 1
    assert result["duplicados"] == 0
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Minha Aba"]
    assert ws.cell(2, 1).value == "João Silva"
    assert ws.cell(2, 2).value == "Acme Corp"
    wb.close()